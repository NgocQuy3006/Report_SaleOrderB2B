# -*- coding: utf-8 -*-

import io
from html import escape

from odoo import http, _
from odoo.http import request

# Tự động nhận diện thư viện tạo file Excel chuyên nghiệp
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None


class BisproSaleOrderReportController(http.Controller):
    """HTML/PDF/XLSX delivery endpoints for Bispro Sale Order report."""

    def _get_so(self, so_id):
        so = request.env["sale.order"].browse(int(so_id)).exists()
        if not so:
            return None

        if hasattr(so, "check_access"):
            so.check_access("read")
        else:
            so.check_access_rights("read")
            so.check_access_rule("read")
        return so

    def _txt(self, value):
        if value in (False, None):
            return ""
        return escape(str(value))

    def _date(self, value):
        if not value:
            return ""
        try:
            return escape(value.strftime("%d/%m/%Y"))
        except Exception:
            return escape(str(value))

    def _money(self, amount, currency):
        symbol = ""
        position = "after"
        if currency:
            symbol = (
                         currency.symbol if "symbol" in currency._fields else ""
                     ) or (
                         currency.name if "name" in currency._fields else ""
                     ) or ""
            position = (
                           currency.position if "position" in currency._fields else "after"
                       ) or "after"

        try:
            text = "{:,.2f}".format(amount or 0.0)
        except Exception:
            text = str(amount or "")

        if position == "before":
            return escape("%s %s" % (symbol, text)).strip()
        return escape("%s %s" % (text, symbol)).strip()

    def _partner_address(self, partner):
        if not partner:
            return ""
        parts = []
        for field_name in ["street", "street2", "city", "state_id", "country_id"]:
            if field_name not in partner._fields:
                continue
            val = partner[field_name]
            if not val:
                continue
            if hasattr(val, "display_name"):
                val = val.display_name
            parts.append(str(val))
        return ", ".join(parts)

    def _field_name(self, record, field_name):
        if not record:
            return ""
        if field_name not in record._fields:
            return ""
        value = record[field_name]
        if hasattr(value, "display_name"):
            return value.display_name or value.name or ""
        return value or ""

    def _line_values(self, line):
        if hasattr(line, "_bispro_line_report_values"):
            return line._bispro_line_report_values()
        return {
            "display_type": getattr(line, "display_type", False),
            "product_code": getattr(getattr(line, "product_id", False), "default_code", "") or "",
            "description": getattr(line, "name", "") or "",
            "uom": getattr(getattr(line, "product_uom", False), "name", "") or "",
            "qty": getattr(line, "product_uom_qty", 0.0) or 0.0,
            "price_unit": getattr(line, "price_unit", 0.0) or 0.0,
            "taxes": ", ".join(line.tax_id.mapped("name")) if hasattr(line, "tax_id") else "",
            "subtotal": getattr(line, "price_subtotal", 0.0) or 0.0,
        }

    # =========================================================================
    # HÀM XUẤT EXCEL THEO THƯ VIỆN OPENPYXL (FIX TRIỆT ĐỂ LỖI KÝ TỰ RÁC)
    # =========================================================================
    @http.route(['/bispro/so/<int:order_id>/xlsx'], type='http', auth='user')
    def bispro_so_xlsx(self, order_id, **kwargs):
        so = self._get_so(order_id)
        if not so:
            return request.not_found()

        if not openpyxl:
            return request.make_response(
                "Hệ thống thiếu thư viện 'openpyxl'. Vui lòng cài đặt bằng lệnh: pip install openpyxl", status=500)

        # Khởi tạo file Excel mới
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sale Order"
        ws.views.sheetView[0].showGridLines = True  # Hiển thị lưới ô Excel

        # Thiết lập màu sắc và font chữ chuẩn phom giống mẫu của bạn
        font_title = Font(name="Arial", size=18, bold=True, color="1F4E78")
        font_header = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        font_sub_header = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        font_bold = Font(name="Arial", size=10, bold=True)
        font_normal = Font(name="Arial", size=10)

        fill_dark = PatternFill(start_color="0A192F", end_color="0A192F", fill_type="solid")
        fill_blue = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        fill_light = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

        align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        align_right = Alignment(horizontal="right", vertical="center")

        thin_side = Side(border_style="thin", color="D9D9D9")
        border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

        # 1. TIÊU ĐỀ LỚN
        ws.merge_cells("A2:H2")
        ws["A2"] = "SALE ORDER"
        ws["A2"].font = font_title
        ws["A2"].alignment = align_right

        # 2. THÔNG TIN ĐƠN HÀNG VÀ CÔNG TY CHỦ QUẢN
        ws["A4"] = so.company_id.name or "My Company"
        ws["A4"].font = font_bold
        ws.merge_cells("F4:G4")
        ws["F4"] = f"SO: {so.name or ''}"
        ws["F4"].font = font_header
        ws["F4"].fill = fill_blue
        ws["F4"].alignment = align_center

        ws["A5"] = self._partner_address(so.company_id.partner_id)
        ws["A5"].font = font_normal
        ws.merge_cells("F5:G5")
        ws["F5"] = "Order Date:"
        ws["F5"].font = font_normal
        ws["H5"] = so.date_order.strftime("%d/%m/%Y") if so.date_order else ""
        ws["H5"].font = font_normal

        ws["A6"] = f"Tax ID: {so.company_id.vat or ''}"
        ws["A6"].font = font_normal
        ws.merge_cells("F6:G6")
        ws["F6"] = "Status:"
        ws["F6"].font = font_normal
        ws["H6"] = dict(so._fields['state'].selection).get(so.state, so.state) if so.state else ""
        ws["H6"].font = font_normal

        # 3. PHÂN ĐOẠN 1. DOCUMENT CONTROL & SHIP TO
        ws.merge_cells("A9:D9")
        ws["A9"] = "1. DOCUMENT CONTROL"
        ws["A9"].font = font_sub_header
        ws["A9"].fill = fill_dark

        ws.merge_cells("E9:H9")
        ws["E9"] = "3. SHIP TO / COMPANY INFORMATION"
        ws["E9"].font = font_sub_header
        ws["E9"].fill = fill_dark

        info_rows = [
            ("SO Numb", so.name or "", "Company", so.company_id.name or ""),
            ("Currency", so.currency_id.name or "", "Address", self._partner_address(so.company_id.partner_id)),
            ("Delivery", "", "Tax ID", so.company_id.vat or ""),
            ("Salesperson", so.user_id.name or "", "Phone/Email",
             f"{so.company_id.phone or ''} / {so.company_id.email or ''}"),
            ("Payment Terms", self._field_name(so, "payment_term_id"), "", "")
        ]

        curr_row = 10
        for r_data in info_rows:
            ws.cell(row=curr_row, column=1, value=r_data[0]).font = font_bold
            ws.cell(row=curr_row, column=2, value=r_data[1]).font = font_normal
            ws.cell(row=curr_row, column=5, value=r_data[2]).font = font_bold
            ws.cell(row=curr_row, column=6, value=r_data[3]).font = font_normal
            ws.merge_cells(start_row=curr_row, start_column=2, end_row=curr_row, end_column=4)
            ws.merge_cells(start_row=curr_row, start_column=6, end_row=curr_row, end_column=8)
            curr_row += 1

        # PHÂN ĐOẠN 2. CUSTOMER INFORMATION
        ws.merge_cells("A15:D15")
        ws["A15"] = "2. CUSTOMER INFORMATION"
        ws["A15"].font = font_sub_header
        ws["A15"].fill = fill_dark

        cust_rows = [
            ("Customer", so.partner_id.name or ""),
            ("Address", self._partner_address(so.partner_id)),
            ("Tax ID", so.partner_id.vat or ""),
            ("Phone/Email", f"{so.partner_id.phone or ''} / {so.partner_id.email or ''}")
        ]
        curr_row = 16
        for c_data in cust_rows:
            ws.cell(row=curr_row, column=1, value=c_data[0]).font = font_bold
            ws.cell(row=curr_row, column=2, value=c_data[1]).font = font_normal
            ws.merge_cells(start_row=curr_row, start_column=2, end_row=curr_row, end_column=4)
            curr_row += 1

        # 4. BẢNG CHI TIẾT DANH SÁCH SẢN PHẨM (PRODUCT LINES)
        prod_start_row = 21
        headers = ["No.", "Product Code", "Description", "UoM", "Qty", "Unit Price", "Taxes", "Subtotal"]
        for col_idx, text in enumerate(headers, 1):
            cell = ws.cell(row=prod_start_row, column=col_idx, value=text)
            cell.font = font_header
            cell.fill = fill_blue
            cell.alignment = align_center
            cell.border = border_all

        idx = 1
        curr_row = prod_start_row + 1
        for line in so.order_line:
            vals = self._line_values(line)
            if vals["display_type"]:
                continue

            ws.cell(row=curr_row, column=1, value=idx).alignment = align_center
            ws.cell(row=curr_row, column=2, value=vals["product_code"]).alignment = align_left
            ws.cell(row=curr_row, column=3, value=vals["description"]).alignment = align_left
            ws.cell(row=curr_row, column=4, value=vals["uom"]).alignment = align_center
            ws.cell(row=curr_row, column=5, value=vals["qty"]).alignment = align_right
            ws.cell(row=curr_row, column=6, value=vals["price_unit"]).alignment = align_right
            ws.cell(row=curr_row, column=7, value=vals["taxes"]).alignment = align_center
            ws.cell(row=curr_row, column=8, value=vals["subtotal"]).alignment = align_right

            # Format dấu phẩy phần nghìn cho tiền tệ số học
            ws.cell(row=curr_row, column=5).number_format = '#,##0.00'
            ws.cell(row=curr_row, column=6).number_format = '#,##0.00'
            ws.cell(row=curr_row, column=8).number_format = '#,##0.00'

            for col_idx in range(1, 9):
                c = ws.cell(row=curr_row, column=col_idx)
                c.font = font_normal
                c.border = border_all

            curr_row += 1
            idx += 1

        if idx == 1:
            ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=8)
            ws.cell(row=curr_row, column=1, value="No order lines").alignment = align_center
            ws.cell(row=curr_row, column=1).font = font_normal
            curr_row += 1

        # 5. KHỐI TỔNG TIỀN ĐƠN HÀNG (TOTALS)
        totals = [
            ("Untaxed Amount", so.amount_untaxed),
            ("Taxes", so.amount_tax),
            ("TOTAL", so.amount_total)
        ]

        for t_label, t_val in totals:
            ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=6)
            lbl_cell = ws.cell(row=curr_row, column=6, value=t_label)
            lbl_cell.font = font_bold if t_label == "TOTAL" else font_normal
            lbl_cell.alignment = align_right

            val_cell = ws.cell(row=curr_row, column=7, value=t_val)
            val_cell.font = font_bold if t_label == "TOTAL" else font_normal
            val_cell.alignment = align_right
            val_cell.number_format = '#,##0.00'

            ws.merge_cells(start_row=curr_row, start_column=7, end_row=curr_row, end_column=8)
            if t_label == "TOTAL":
                lbl_cell.fill = fill_light
                val_cell.fill = fill_light

            curr_row += 1

        # 6. ĐIỀU KHOẢN KÈM THEO (TERMS & CONDITIONS)
        curr_row += 1
        ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=8)
        ws.cell(row=curr_row, column=1, value="4. TERMS, CONDITIONS & NOTES").font = font_sub_header
        ws.cell(row=curr_row, column=1).fill = fill_dark
        curr_row += 1

        ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=8)
        note_text = so.note or "Please supply goods/services according to this Sales Order, agreed commercial terms, applicable tax regulations, quality standards, and delivery schedule."
        ws.cell(row=curr_row, column=1, value=note_text).font = font_normal
        ws.cell(row=curr_row, column=1).alignment = align_left

        # 7. TỰ ĐỘNG CĂN RỘNG KÍCH THƯỚC CỘT THEO CHIỀU DÀI CHỮ
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value and not isinstance(cell.value, int) and str(cell.value).startswith('='):
                    continue
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        # Chuyển đổi dữ liệu ra luồng byte
        fp = io.BytesIO()
        wb.save(fp)
        excel_data = fp.getvalue()
        fp.close()

        return request.make_response(
            excel_data,
            headers=[
                ('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                ('Content-Disposition', f'attachment; filename={so.name or "SaleOrder"}.xlsx')
            ]
        )